#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Author : ALEXuH
# @Created on : 2019/9/14 0014

from openpyxl import load_workbook
import requests
from os import path as _path

def unify(v):
    if v is None:
        return v
    if isinstance(v, unicode):
        return v.encode("utf8")
    if isinstance(v, str):
        return v
    return str(v)

def getCol():
    map = {}
    str = "I J K L M N O P Q R S T U V W X Y Z AA AB AC AD AE AF AG AH AI AJ AK AL AM"
    list = str.split(" ")
    print(len(list))
    for i in range(0, 31):
        map[i+1] = list[i]
    return map

import configparser
import os

cf = configparser.ConfigParser()
config_path = _path.join(_path.dirname(os.path.realpath(__file__)), "source.ini")
print(config_path)
#cf.read_file()
cf.read(config_path, encoding="utf-8")
options = cf.options("Excel")
file1 = _path.join(_path.dirname(os.path.realpath(__file__)), cf.get("Excel", "file1"))
file2 = _path.join(_path.dirname(os.path.realpath(__file__)), cf.get("Excel", "file2"))
file3 = _path.join(_path.dirname(os.path.realpath(__file__)), cf.get("Excel", "file3"))
number = int(cf.get("Excel", "number"))
special_event = cf.get("Excel", "special_event")
passenger_flow = cf.get("Excel", "passenger_flow")
file_name = cf.get("Excel", "file_name")
file_info_name = cf.get("Excel", "file_info_name")
percent = cf.get("Excel", "percent")
#sale_default = cf.get("Excel", "sale_default")
over_sale = float(cf.get("Excel", "over_sale"))
diff_value = float(cf.get("Excel", "diff_value"))
diff_file = cf.get("Excel", "diff_file")

path = file1
wb = load_workbook(file1)
ws = wb.get_sheet_by_name(unicode("销售", "utf-8"))
map_date = getCol()
print(map_date)
c = map_date[number]
y = map_date[number-1]
map = {}
a = 0
for row in range(6, 382):
    cell = str(row)
    g = "G" + str(cell)
    if unify(ws[g].value) is not None:
        value = unify(ws[g].value).strip().upper()
        if value not in map:
            a += 1
            map[value] = c+cell
import xlrd
map_mian = {}
sheet = xlrd.open_workbook(file2)
for page in range(0,1):
    table = sheet.sheet_by_index(page)
    for i in range(1, table.nrows):
        row = table.row_values(i)
        merchant = unify(row[2]).strip().upper()
        sales = unify(row[3]).strip().upper()
        map_mian[merchant] = sales
#print(len(map_mian.keys()))

map_no_mian = {}
sheet2 = xlrd.open_workbook(file3)
for page in range(0,1):
    table = sheet2.sheet_by_index(page)
    for i in range(1, table.nrows):
        row = table.row_values(i)
        merchant = unify(row[1]).strip().upper()
        sales = unify(row[2]).strip().upper()
        print(row[1])
        print(row[2])
        map_no_mian[merchant] = sales
#print(len(map_no_mian.keys()))

error = []
for k,v in map_mian.items():
    print(float(v))
    if k not in map_no_mian and float(v) == 0.0:
        error.append(k)
diff = []
fl = open(diff_file, "w+")
for k,v in map_mian.items():
    sale1 = float(map_mian[k])
    if k not in map_no_mian and sale1 == 0:
        map_mian.pop(k)
for k,v in map_no_mian.items():
    if k not in map_mian:
        map_mian[k] = float(v)
    else:
        sale1 = float(map_mian[k])
        sale2 = float(map_no_mian[k])
        fl.write(k+":"+str(int(sale1))+"|"+str(int(sale2)))
        fl.write("\n")
        # if sale1 >= sale2:
        #     map_mian[k] = sale1
        # else:
        #     map_mian[k] = sale2
        if sale1 == 0:
            map_mian[k] = sale2
        if abs(sale1 - sale2) >= diff_value:
            diff.append(k)
fl.close()
print(map_mian)
print(map)
k1 = []
for k,v in map.items():
    if k in map_mian:
        sale1 = float(map_mian[k])
        pos = map[k]
        if int(sale1) == -1:
            ws[pos] = "装修"
        elif int(sale1) == -2:
            ws[pos] = "未进场"
        else:
            ws[pos] = sale1
    else:
        pos = map[k]
        k1.append(k)
        #print("not in: " + k)

url = "http://wthrcdn.etouch.cn/weather_mini?city=常州"
res = requests.get(url)
weather = res.json()["data"]["yesterday"]["type"]
ws[c+"3"] = weather
ws[c+"4"] = float(passenger_flow)
ws[c+"5"] = special_event

print(weather)
print(len(k1))
print(len(map.keys()))
print(len(map_mian.keys()))
wb.save(file_name)
# 计算over_percent的品牌
over_percent = []
over_sale_topic = []
for row in range(6, 382):
    cell = str(row)
    g = "G" + cell
    yes = y + cell
    today = c + cell
    try:
        v1 = float(ws[today].value)
        topic = unify(ws[g].value)
        if v1 >= over_sale:
            over_sale_topic.append(topic)
        v2 = float(ws[yes].value)
        if (v1 - v2) / v2 >= float(percent):
            over_percent.append(topic)
    except:
        pass
    if unify(ws[g].value) is not None:
        value = unify(ws[g].value).strip().upper()
        if value not in map:
            print(value)
            a += 1
            map[value] = c+cell
k2 = []
for k,v in map_mian.items():
    if k not in map:
        k2.append(k)

import sys
from importlib import reload
reload(sys)
sys.setdefaultencoding('utf8')
with open(file_info_name, "w+") as f:
    f.write("天气:" + weather)
    f.write("\n")
    f.write("客流量: "+ passenger_flow)
    f.write("\n")
    f.write("特殊事件: "+ special_event)
    f.write("\n")
    f.write("\n")
    f.write("读取最终表商家总数: " + str(len(map.keys())))
    f.write("\n")
    f.write("其他两张表商家总数: " + str(len(map_mian.keys())))
    f.write("\n")
    f.write("\n")
    f.write("未匹配上的商家数量: {0}".format(len(k1)))
    f.write("\n")
    f.write("未匹配上的商家详细信息: {0}".format(",".join(k1)))
    f.write("\n")
    f.write("\n")
    f.write("系统表为0且商家自己上报表中无数据的数量: {0}".format(len(error)))
    f.write("\n")
    f.write("系统表为0且商家自己上报表中无数据的商家信息: {0}".format(",".join(error)))
    f.write("\n")
    f.write("\n")
    f.write("手工表和系统表中有的且最终输出表中没有的的商家数量: {1}".format(percent, len(k2)))
    f.write("\n")
    f.write("手工表和系统表中有的且最终输出表中没有的商家详细信息:{1}".format(percent, ",".join(k2)))
    f.write("\n")
    f.write("\n")
    f.write("每日增长额度超过{0}的商家数量: {1}".format(percent, len(over_percent)))
    f.write("\n")
    f.write("每日增长额度超过{0}的商家详细信息:{1}".format(percent, ",".join(over_percent)))
    f.write("\n")
    f.write("\n")
    f.write("两张表中差值超过{0}的商家数量: {1}".format(diff_value, len(diff)))
    f.write("\n")
    f.write("两张表中差值超过{0}的商家详细信息:{1}".format(diff_value, ",".join(diff)))
    f.write("\n")
    f.write("\n")
    f.write("销售额大于{0}的商家数量: {1}".format(over_sale, len(over_sale_topic)))
    f.write("\n")
    f.write("销售额大于{0}的商家的详细信息:{1}".format(over_sale, ",".join(over_sale_topic)))
raw_input()